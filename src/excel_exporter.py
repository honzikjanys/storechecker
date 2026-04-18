"""
Excel export module for StoreChecker
Generates .xlsx files with current store data
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export store data to Excel (.xlsx) files"""
    
    def __init__(self, output_dir: str = "./data/reports"):
        """Initialize ExcelExporter
        
        Args:
            output_dir: Directory where Excel files will be saved
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"Excel exporter initialized, output dir: {self.output_dir}")
    
    def export_all_stores(self, stores: List[Dict], timestamp: Optional[datetime] = None) -> Path:
        """Export all stores to a single Excel file
        
        Args:
            stores: List of store dictionaries from database
            timestamp: Optional timestamp for filename (defaults to now)
            
        Returns:
            Path to generated Excel file
        """
        if not timestamp:
            timestamp = datetime.now()
        
        filename = f"prodejny_{timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Prodejny"
        
        # Add headers
        headers = ["Řetězec", "Název", "Město", "Adresa", "Status", "Telefon", "Hodiny", "Poslední aktualizace"]
        ws.append(headers)
        
        # Style headers
        self._style_header(ws, headers)
        
        # Add store data
        for store in stores:
            row = [
                store.get('retailer', ''),
                store.get('name', ''),
                store.get('city', ''),
                store.get('address', ''),
                store.get('status', ''),
                store.get('phone', ''),
                store.get('hours', ''),
                store.get('updated_at', '')
            ]
            ws.append(row)
            
            # Color code status
            self._color_status_cell(ws, len(ws.rows), store.get('status', ''))
        
        # Adjust column widths
        self._adjust_columns(ws, headers)
        
        # Save workbook
        wb.save(filepath)
        logger.info(f"Exported {len(stores)} stores to {filepath}")
        
        return filepath
    
    def export_by_retailer(self, stores_by_retailer: Dict[str, List[Dict]], 
                          timestamp: Optional[datetime] = None) -> Dict[str, Path]:
        """Export stores grouped by retailer (one sheet per retailer)
        
        Args:
            stores_by_retailer: Dictionary with retailer names as keys and store lists as values
            timestamp: Optional timestamp for filename
            
        Returns:
            Dictionary mapping retailer names to exported file paths
        """
        if not timestamp:
            timestamp = datetime.now()
        
        filename = f"prodejny_podle_retezcu_{timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        
        wb = Workbook()
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        headers = ["Název", "Město", "Adresa", "Status", "Telefon", "Hodiny", "Poslední aktualizace"]
        
        # Create sheet for each retailer
        for retailer, stores in stores_by_retailer.items():
            # Sanitize sheet name (max 31 chars, no special chars)
            sheet_name = str(retailer).replace('/', '-')[:31]
            ws = wb.create_sheet(sheet_name)
            
            # Add headers
            ws.append(headers)
            self._style_header(ws, headers)
            
            # Add store data
            for store in stores:
                row = [
                    store.get('name', ''),
                    store.get('city', ''),
                    store.get('address', ''),
                    store.get('status', ''),
                    store.get('phone', ''),
                    store.get('hours', ''),
                    store.get('updated_at', '')
                ]
                ws.append(row)
                self._color_status_cell(ws, len(ws.rows), store.get('status', ''))
            
            # Adjust columns
            self._adjust_columns(ws, headers)
            
            logger.info(f"  {retailer}: {len(stores)} stores")
        
        # Save workbook
        wb.save(filepath)
        logger.info(f"Exported stores by retailer to {filepath}")
        
        return {retailer: filepath for retailer in stores_by_retailer.keys()}
    
    def export_changes(self, changes: Dict[str, List[Dict]], timestamp: Optional[datetime] = None) -> Path:
        """Export detected changes to Excel file
        
        Args:
            changes: Dictionary with change types as keys ('new', 'temporarily_closed', etc.)
            timestamp: Optional timestamp for filename
            
        Returns:
            Path to generated Excel file
        """
        if not timestamp:
            timestamp = datetime.now()
        
        filename = f"zmeny_{timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Změny"
        
        # Add summary
        ws.append(["Souhrn detekovaných změn"])
        ws['A1'].font = Font(size=14, bold=True)
        ws.append([])  # Empty row
        
        # Add stats
        ws.append(["Typ změny", "Počet"])
        stats_headers = ["Typ změny", "Počet"]
        self._style_header(ws, stats_headers)
        
        change_types = {
            'new': 'Nové prodejny',
            'temporarily_closed': 'Dočasně zavřeno',
            'permanently_closed': 'Trvale zavřeno',
            'reopened': 'Znovuotevřeno'
        }
        
        row_num = 3
        for change_key, change_label in change_types.items():
            count = len(changes.get(change_key, []))
            ws.append([change_label, count])
            row_num += 1
        
        ws.append([])  # Empty row
        
        # Add details for each change type
        row_num = row_num + 2
        headers = ["Řetězec", "Název", "Město", "Adresa", "Typ změny", "Datum"]
        
        for change_type, change_label in change_types.items():
            items = changes.get(change_type, [])
            if not items:
                continue
            
            # Section header
            ws[f'A{row_num}'] = change_label
            ws[f'A{row_num}'].font = Font(bold=True, size=12)
            row_num += 1
            
            # Column headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.value = header
                self._style_header_cell(cell)
            row_num += 1
            
            # Data rows
            for item in items:
                row = [
                    item.get('retailer', ''),
                    item.get('store_id', '').split('_')[1] if '_' in item.get('store_id', '') else '',
                    '',  # City not in history
                    '',  # Address not in history
                    change_type,
                    item.get('change_date', '')
                ]
                ws.append(row)
                row_num += 1
        
        # Adjust columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        
        wb.save(filepath)
        logger.info(f"Exported changes to {filepath}")
        
        return filepath
    
    def _style_header(self, ws, headers: List[str]) -> None:
        """Style header row"""
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            self._style_header_cell(cell)
    
    def _style_header_cell(self, cell) -> None:
        """Style a single header cell"""
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    def _color_status_cell(self, ws, row: int, status: str) -> None:
        """Color code status cell based on status value"""
        cell = ws.cell(row=row, column=5)  # Status is the 5th column
        
        colors = {
            'open': 'C6EFCE',  # Green
            'temporarily_closed': 'FFEB9C',  # Yellow
            'permanently_closed': 'FFC7CE'  # Red
        }
        
        color = colors.get(status, 'FFFFFF')  # White default
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    def _adjust_columns(self, ws, headers: List[str]) -> None:
        """Auto-adjust column widths based on header length"""
        column_widths = {
            "Řetězec": 15,
            "Název": 30,
            "Město": 20,
            "Adresa": 30,
            "Status": 18,
            "Telefon": 15,
            "Hodiny": 20,
            "Poslední aktualizace": 20,
            "Typ změny": 20,
            "Počet": 10,
            "Datum": 15
        }
        
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            width = column_widths.get(header, 20)
            ws.column_dimensions[col_letter].width = width
